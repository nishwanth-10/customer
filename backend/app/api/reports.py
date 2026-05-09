"""
Reports API — PDF and Excel export for payments, transactions, and customer history.
"""
from fastapi import APIRouter, Depends, Query, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from datetime import date
import uuid
import io

from app.core.database import get_db
from app.core.deps import get_current_active_user
from app.models.customers import Customer
from app.models.transactions import Transaction, TransactionType
from app.models.payments import Payment
from app.models.users import User

router = APIRouter(prefix="/reports", tags=["Reports"])


@router.get("/monthly-summary")
async def monthly_summary(
    business_id: uuid.UUID = Query(...),
    month: int = Query(..., ge=1, le=12),
    year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Get monthly payment summary with per-customer breakdown."""
    import calendar
    _, last_day = calendar.monthrange(year, month)
    start_date = date(year, month, 1)
    end_date = date(year, month, last_day)

    # All customers in this business
    customers_result = await db.execute(
        select(Customer).where(Customer.business_id == business_id, Customer.is_active == True)
    )
    customers = customers_result.scalars().all()

    summary_items = []
    total_collected = 0
    total_pending = 0

    for customer in customers:
        # Collections for this month
        collected = (await db.execute(
            select(func.sum(Transaction.amount)).where(
                Transaction.customer_id == customer.id,
                Transaction.transaction_type == TransactionType.CREDIT,
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date,
            )
        )).scalar() or 0

        due = float(customer.monthly_payment_amount) - float(collected)
        total_collected += float(collected)
        total_pending += max(0, due)

        summary_items.append({
            "customer_id": str(customer.id),
            "name": customer.name,
            "mobile": customer.mobile_number,
            "monthly_due": float(customer.monthly_payment_amount),
            "collected": float(collected),
            "pending": max(0, due),
            "status": "paid" if collected >= customer.monthly_payment_amount else ("partial" if collected > 0 else "unpaid"),
        })

    return {
        "month": month,
        "year": year,
        "total_customers": len(customers),
        "paid_customers": sum(1 for s in summary_items if s["status"] == "paid"),
        "total_collected": total_collected,
        "total_pending": total_pending,
        "items": summary_items,
    }


@router.get("/export/pdf")
async def export_pdf(
    business_id: uuid.UUID = Query(...),
    month: int = Query(..., ge=1, le=12),
    year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export monthly report as PDF."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    # Get summary data
    summary_data = await monthly_summary(business_id, month, year, db, current_user)
    summary_data = summary_data if isinstance(summary_data, dict) else {}

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph(f"Monthly Report — {month}/{year}", styles["Title"]))
    story.append(Spacer(1, 12))

    # Summary stats
    story.append(Paragraph(f"Total Collected: ₹{summary_data.get('total_collected', 0):,.2f}", styles["Normal"]))
    story.append(Paragraph(f"Total Pending: ₹{summary_data.get('total_pending', 0):,.2f}", styles["Normal"]))
    story.append(Spacer(1, 12))

    # Table
    table_data = [["Name", "Mobile", "Monthly Due", "Collected", "Pending", "Status"]]
    for item in summary_data.get("items", []):
        table_data.append([
            item["name"],
            item["mobile"],
            f"₹{item['monthly_due']:,.0f}",
            f"₹{item['collected']:,.0f}",
            f"₹{item['pending']:,.0f}",
            item["status"].upper(),
        ])

    table = Table(table_data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3730A3")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    story.append(table)

    doc.build(story)
    buffer.seek(0)

    return Response(
        content=buffer.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=report_{year}_{month:02d}.pdf"},
    )


@router.get("/export/excel")
async def export_excel(
    business_id: uuid.UUID = Query(...),
    month: int = Query(..., ge=1, le=12),
    year: int = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export monthly report as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    summary_data = await monthly_summary(business_id, month, year, db, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Report {month}-{year}"

    # Header row
    headers = ["Name", "Mobile", "Monthly Due (₹)", "Collected (₹)", "Pending (₹)", "Status"]
    header_fill = PatternFill(start_color="3730A3", end_color="3730A3", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, item in enumerate(summary_data.get("items", []), 2):
        ws.append([
            item["name"],
            item["mobile"],
            item["monthly_due"],
            item["collected"],
            item["pending"],
            item["status"].upper(),
        ])

    # Summary row
    ws.append([])
    ws.append(["TOTAL", "", "", summary_data.get("total_collected", 0), summary_data.get("total_pending", 0), ""])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=report_{year}_{month:02d}.xlsx"},
    )
